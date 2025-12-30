from openpyxl import load_workbook

# Load Excel workbook
wb = load_workbook("Student_Performance_Analytics.xlsx")

# Select sheets
data_sheet = wb["Student_Data"]
analytics_sheet = wb["Analytics"]

students = []

# Read student data (skip header)
for row in data_sheet.iter_rows(min_row=2, values_only=True):
    roll, name, m1, m2, m3, _, _, _, _ = row
    marks = [m1, m2, m3]

    total = sum(marks)
    average = total / len(marks)

    if average >= 90:
        grade = "A"
    elif average >= 75:
        grade = "B"
    elif average >= 60:
        grade = "C"
    Else:
        grade = "Fail"

    status = "PASS" if min(marks) >= 35 else "FAIL"

    students.append({
        "roll": roll,
        "name": name,
        "total": total,
        "average": average,
        "grade": grade,
        "status": status
    })

# Write computed values back to Excel
row_num = 2
for s in students:
    data_sheet[f"F{row_num}"] = s["total"]
    data_sheet[f"G{row_num}"] = s["average"]
    data_sheet[f"H{row_num}"] = s["grade"]
    data_sheet[f"I{row_num}"] = s["status"]
    row_num += 1

# Class analytics
class_avg = sum(s["average"] for s in students) / len(students)
topper = max(students, key=lambda x: x["average"])

analytics_sheet["B3"] = class_avg
analytics_sheet["B4"] = topper["average"]
analytics_sheet["B5"] = topper["name"]

# Save workbook
wb.save("Student_Performance_Analytics.xlsx")

print("Excel Analytics Updated Successfully!")
print("Class Average:", round(class_avg, 2))
print("Topper:", topper["name"])
